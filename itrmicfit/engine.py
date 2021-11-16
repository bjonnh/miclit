from .plotting import plot
from .sourcedata import SourceData
from .fitting import fit_from_sourcedata, Fit
from openpyxl.utils.dataframe import dataframe_to_rows

import io

import openpyxl
import openpyxl.drawing.image
import pandas as pd
from PIL import Image
from typing import List

import matplotlib.pyplot as plt


class LoadError(Exception):
    """An error raised when the loading of an initial file failed"""
    pass


class Engine:
    data = None
    fits: List[Fit] = []
    MICn: int = 90
    DEFAULT_COLUMNS = ["blank", '0', '1', '2', '3', '4', '5', '6', '7', '8',
                       "bacterial_control", "blank2", "name", "initial_concentration"]

    onLoadDone = None  # callback for when loading is done
    onFitDone = None  # callback for when the fitting is done
    onExportDone = None  # call back for when export is done

    def __init__(self):
        self.data = None

    def load_file(self, file):
        self.clean_fits()

        # That's the way we detect "raw" files

        file = openpyxl.open(file)

        worksheets_names = [name.title for name in file.worksheets]

        if "End point" in worksheets_names:
            source_df = pd.read_excel(file, usecols="B:O", skiprows=15, header=None, engine="openpyxl",
                                      sheet_name="End point").dropna(axis=0, how="all")

        elif "RAW" in worksheets_names:
            source_df = pd.read_excel(file, usecols="A:N", skiprows=0, header=None, engine="openpyxl",
                                      sheet_name="RAW").dropna(axis=0, how="all")
        else:
            raise LoadError(
                """The loaded file does not contain a file in one of the supported formats.
                If you made a combined file, did you call the sheet "RAW"?
                """)
        default_columns = self.DEFAULT_COLUMNS

        # Do we have a file with no name and initial concentrations?
        if len(source_df.columns) == len(default_columns) - 2:
            source_df["name"] = [""] * len(source_df)
            source_df["initial_concentration"] = [1] * len(source_df)
        source_df.columns = default_columns

        if self.data is not None:
            new_data = pd.concat([self.data.data, source_df]).reset_index(drop=True)
            self.data = SourceData(new_data)
        else:
            self.data = SourceData(source_df.reset_index(drop=True))

        if self.onLoadDone is not None:
            self.onLoadDone()

    def clear(self):
        """Clear the current data"""
        self.clean_fits()
        self.data = None

    def clean_fits(self):
        self.fits = []

    def fit(self):
        self.data.calculate()
        self.fits = fit_from_sourcedata(self.data, self.MICn / 100.0)

        if self.onFitDone is not None:
            self.onFitDone()

    def export_as_dataframe(self):
        data = []

        for index, fit in enumerate(self.fits):
            data.append({"name": fit.name,
                         "initial concentration": self.data.initial_concentrations[index],
                         "type of fit": fit.type_of_fit.value,
                         "mic percentage": int(fit.mic.percentage * 100),
                         "mic concentration": str(fit.mic.concentration),
                         "mic quality": fit.mic.quality.value,
                         "mic uncertainty": str(fit.uncertainties["mic"]),
                         "hill4p_d0i": fit.fitted_curve[0][0],
                         "hill4p_n": fit.fitted_curve[0][1],
                         "hill4p_s": fit.fitted_curve[0][2],
                         "hill4p_o": fit.fitted_curve[0][3]
                         })
        return pd.DataFrame(data)

    def export_as_spreadsheet(self, file, graphics=True, fits=None):
        """Export as a spreadsheet both the data and the graphic, with large files,
        we may want to not export the graphics."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            ["name", "initial_concentration", "type_of_fit", "mic_percentage", "mic_concentration", "mic_quality",
             "mic_uncertainty", "hill4p_d0i", "hill4p_n", "hill4p_s", "hill4p_o"])
        for index, fit in enumerate(self.fits):
            ws.append([fit.name,
                       self.data.initial_concentrations[index],
                       fit.type_of_fit.value,
                       int(fit.mic.percentage * 100),
                       fit.mic.concentration,
                       fit.mic.quality.value,
                       str(fit.uncertainties["mic"]),
                       fit.fitted_curve[0][0],
                       fit.fitted_curve[0][1],
                       fit.fitted_curve[0][2],
                       fit.fitted_curve[0][3]
                       ])

        ws_source = wb.create_sheet("RAW")
        for r in dataframe_to_rows(self.data.data, index=False, header=False):
            ws_source.append(r)

        if graphics is True:
            ws1 = wb.create_sheet("Graphic")
            fig = plt.figure(1, (20, 20), tight_layout=True)
            plot(fits, fig=fig)
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=100)
            buf.seek(0)
            im = Image.open(buf)
            img = openpyxl.drawing.image.Image(im)
            img.anchor = "A1"
            ws1.add_image(img)

        wb.save(file)

        if self.onExportDone is not None:
            self.onExportDone(file)
